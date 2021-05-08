# coding:utf-8
# !/usr/bin/python
# ========================================================
# Project: automationTest
# Creator: lilyluo
# Create time: 2021-05-09 06:46
# IDE: PyCharm
# =========================================================

file_folder = "/Users/lilyluo/Documents/project/automationTest/data"
import os
import openpyxl

# create work book
def create_workbook():

    nwb = openpyxl.Workbook()
    for n in range(1,13):

        book_name = os.path.join(file_folder,"{:02}.xlsx".format(n))
        nwb.save(book_name)

#sheet operation
def create_sheet(filename):
    """create sheet"""
    workbook = openpyxl.load_workbook(os.path.join(file_folder, filename)) #read work book
    workbook.create_sheet()
    workbook.create_sheet("salary")
    workbook.create_sheet("Total",0)
    # workbook.create_chartsheet("test")
    # workbook.create_named_range("123")
    workbook.save(os.path.join(file_folder, filename))

def ope_sheet(filename):
    workbook = openpyxl.load_workbook(os.path.join(file_folder, filename))
    print(workbook.worksheets)
    print([ws for ws in workbook.worksheets])

    #copy sheet
    workbook.copy_worksheet(workbook.worksheets[2]).title="salary_copy"
    print(workbook.sheetnames)
    workbook.save(os.path.join(file_folder,filename))


# create_sheet("02.xlsx")
ope_sheet("02.xlsx")

